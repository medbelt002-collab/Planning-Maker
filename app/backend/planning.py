"""Excel engine: reads/writes the Planning d'Occupation while preserving all
formatting (borders, colors, merged cells, comments, column widths, row heights).

An *entry* is one planning row grouping reservations that share the merge key
(Nom / Entité / Ligne budgétaire / Demandeur). Occupancy and notes are keyed by
day-of-month (1..31) because each sheet represents one month.
"""
from __future__ import annotations

import copy
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from openpyxl.cell.cell import MergedCell

from . import config as C


# ---------------------------------------------------------------------------
# Layout detection
# ---------------------------------------------------------------------------
class SheetLayout:
    def __init__(self, ws):
        self.ws = ws
        self.day_cols: Dict[int, int] = {}
        self.last_day = 0
        self.header_rows: List[int] = []
        self.studio_title_row: Optional[int] = None
        self.chambre_title_row: Optional[int] = None
        self.studio_start = 0
        self.studio_end = 0
        self.chambre_start = 0
        self.chambre_end = 0
        self._detect()

    def _is_day_header_row(self, r: int) -> bool:
        a = self.ws.cell(row=r, column=C.FIRST_DAY_COL).value
        b = self.ws.cell(row=r, column=C.FIRST_DAY_COL + 1).value
        try:
            return int(a) == 1 and int(b) == 2
        except (TypeError, ValueError):
            return False

    def _detect(self):
        ws = self.ws
        max_r = ws.max_row
        for r in range(1, max_r + 1):
            if self._is_day_header_row(r):
                self.header_rows.append(r)
        if self.header_rows:
            hr = self.header_rows[0]
            for c in range(C.FIRST_DAY_COL, ws.max_column + 2):
                v = ws.cell(row=hr, column=c).value
                try:
                    d = int(v)
                except (TypeError, ValueError):
                    continue
                if 1 <= d <= 31:
                    self.day_cols[d] = c
                    self.last_day = max(self.last_day, d)

        for r in range(1, max_r + 1):
            row_text = C.normalize(ws.cell(row=r, column=C.FIRST_DAY_COL).value) + " " \
                       + C.normalize(ws.cell(row=r, column=C.COL_NOM).value)
            if self.studio_title_row is None and C.STUDIOS_MARKER in row_text:
                self.studio_title_row = r
            if self.chambre_title_row is None and C.CHAMBRES_MARKER in row_text:
                self.chambre_title_row = r

        if self.studio_title_row:
            self.studio_start = self.studio_title_row + 1
            ch_header = None
            for r in self.header_rows:
                if self.chambre_title_row and r < self.chambre_title_row:
                    ch_header = r
            if ch_header and ch_header > self.studio_start:
                self.studio_end = ch_header - 1
            elif self.chambre_title_row:
                self.studio_end = self.chambre_title_row - 1
            else:
                self.studio_end = max_r
        if self.chambre_title_row:
            self.chambre_start = self.chambre_title_row + 1
            self.chambre_end = max_r

    def col_for(self, day: int) -> Optional[int]:
        return self.day_cols.get(day)

    def range_for(self, section: str) -> Tuple[int, int]:
        if section == "STUDIOS":
            return self.studio_start, self.studio_end
        return self.chambre_start, self.chambre_end


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _capture_row_style(ws, row: int, last_col: int) -> dict:
    styles = {}
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        if cell.has_style:
            styles[c] = {
                "font": copy.copy(cell.font),
                "border": copy.copy(cell.border),
                "fill": copy.copy(cell.fill),
                "number_format": cell.number_format,
                "protection": copy.copy(cell.protection),
                "alignment": copy.copy(cell.alignment),
            }
    return {"styles": styles, "height": ws.row_dimensions[row].height}


def _apply_row_style(ws, row: int, tmpl: dict, last_col: int):
    for c in range(1, last_col + 1):
        st = tmpl["styles"].get(c)
        if not st:
            continue
        cell = ws.cell(row=row, column=c)
        cell.font = copy.copy(st["font"])
        cell.border = copy.copy(st["border"])
        cell.fill = copy.copy(st["fill"])
        cell.number_format = st["number_format"]
        cell.protection = copy.copy(st["protection"])
        cell.alignment = copy.copy(st["alignment"])
    if tmpl["height"]:
        ws.row_dimensions[row].height = tmpl["height"]


def _clear_cell_content(cell):
    if isinstance(cell, MergedCell):
        return
    cell.value = None
    cell.comment = None
    cell.fill = PatternFill(fill_type=None)


def _safe_set(cell, value):
    """Set a cell value, skipping merged (non-anchor) cells."""
    if isinstance(cell, MergedCell):
        return
    cell.value = value


# ---------------------------------------------------------------------------
# Entry (a planning row)
# ---------------------------------------------------------------------------
class Entry:
    def __init__(self, section: str):
        self.section = section
        self.nom = ""
        self.demandeur = ""
        self.ligne_budgetaire = ""
        self.entite = ""
        self.refs: List[str] = []
        self.status_note = ""
        self.occupancy: Dict[int, int] = {}   # day -> persons
        self.notes: Dict[int, str] = {}        # day -> note

    @property
    def refs_text(self) -> str:
        return "/".join(self.refs)


# ---------------------------------------------------------------------------
# Planning workbook wrapper
# ---------------------------------------------------------------------------
class Planning:
    def __init__(self, path):
        self.path = str(path)
        self.wb = openpyxl.load_workbook(self.path)
        self.layouts: Dict[str, SheetLayout] = {}
        for ws in self.wb.worksheets:
            self.layouts[ws.title] = SheetLayout(ws)

    # -- sheet resolution ---------------------------------------------------
    def sheet_for_month(self, month: int, create=True):
        for ws in self.wb.worksheets:
            if C.month_matches(ws.title, month):
                return ws
        if not create:
            return None
        template_ws = self.wb.worksheets[0]
        new_ws = self.wb.copy_worksheet(template_ws)
        new_ws.title = C.MONTHS_FR[month]
        lay = SheetLayout(new_ws)
        self.layouts[new_ws.title] = lay
        self._clear_data_region(new_ws, lay)
        return new_ws

    def layout(self, ws) -> SheetLayout:
        lay = self.layouts.get(ws.title)
        if lay is None:
            lay = SheetLayout(ws)
            self.layouts[ws.title] = lay
        return lay

    def sheet_month(self, ws) -> Optional[int]:
        for m in range(1, 13):
            if C.month_matches(ws.title, m):
                return m
        return None

    # -- clearing -----------------------------------------------------------
    def _clear_data_region(self, ws, lay: SheetLayout):
        last_col = max(lay.day_cols.values()) if lay.day_cols else 37
        for (start, end) in ((lay.studio_start, lay.studio_end),
                             (lay.chambre_start, lay.chambre_end)):
            if not start:
                continue
            # unmerge any cell merged across the day columns inside the data
            # region (e.g. a whole row merged into one cell) so every day keeps
            # its own individual cell, consistent with the other rows.
            for mc in list(ws.merged_cells.ranges):
                if mc.min_row >= start and mc.max_row <= end:
                    ws.unmerge_cells(str(mc))
            for r in range(start, end + 1):
                for c in range(1, last_col + 1):
                    _clear_cell_content(ws.cell(row=r, column=c))

    # -- import existing rows ----------------------------------------------
    def read_existing_entries(self) -> List[dict]:
        out = []
        for ws in self.wb.worksheets:
            lay = self.layout(ws)
            month = self.sheet_month(ws)
            if not month:
                continue
            for section in ("STUDIOS", "CHAMBRES"):
                start, end = lay.range_for(section)
                if not start:
                    continue
                for r in range(start, end + 1):
                    nom = ws.cell(row=r, column=C.COL_NOM).value
                    ref = ws.cell(row=r, column=C.COL_REF).value
                    if not nom and not ref:
                        continue
                    occ = {}
                    for day, col in lay.day_cols.items():
                        v = ws.cell(row=r, column=col).value
                        if v not in (None, ""):
                            try:
                                occ[day] = int(v)
                            except (TypeError, ValueError):
                                pass
                    out.append({
                        "section": section, "sheet": ws.title, "month": month, "row": r,
                        "nom": nom or "",
                        "refs": [x.strip() for x in str(ref or "").split("/") if x.strip()],
                        "demandeur": ws.cell(row=r, column=C.COL_DEMANDEUR).value or "",
                        "ligne_budgetaire": ws.cell(row=r, column=C.COL_LBUDG).value or "",
                        "entite": ws.cell(row=r, column=C.COL_ENTITE).value or "",
                        "occupancy": occ,
                    })
        return out

    # -- rendering ----------------------------------------------------------
    def render(self, entries_by_month: Dict[int, Dict[str, List[Entry]]]):
        """entries_by_month: {month: {"STUDIOS":[Entry], "CHAMBRES":[Entry]}}
        Each Entry.occupancy/notes are keyed by day-of-month for that month."""
        for month, sections in entries_by_month.items():
            ws = self.sheet_for_month(month, create=True)
            lay = self.layout(ws)
            self._clear_data_region(ws, lay)
            last_col = max(lay.day_cols.values()) if lay.day_cols else 37

            self._write_section(ws, lay, "STUDIOS", sections.get("STUDIOS", []), last_col)
            lay = SheetLayout(ws)               # recompute after inserts
            self.layouts[ws.title] = lay
            self._write_section(ws, lay, "CHAMBRES", sections.get("CHAMBRES", []), last_col)

    def _write_section(self, ws, lay: SheetLayout, section: str,
                       entries: List[Entry], last_col: int):
        start, end = lay.range_for(section)
        if not start:
            return
        capacity = end - start + 1
        tmpl = _capture_row_style(ws, start, last_col)

        if len(entries) > capacity:
            extra = len(entries) - capacity
            ws.insert_rows(end + 1, extra)
            for i in range(extra):
                _apply_row_style(ws, end + 1 + i, tmpl, last_col)

        r = start
        for entry in entries:
            self._write_entry_row(ws, lay, r, entry, tmpl, last_col)
            r += 1

    def _write_entry_row(self, ws, lay: SheetLayout, r: int, entry: Entry,
                         tmpl: dict, last_col: int):
        _apply_row_style(ws, r, tmpl, last_col)
        _safe_set(ws.cell(row=r, column=C.COL_NOM), entry.nom)
        _safe_set(ws.cell(row=r, column=C.COL_REF), entry.refs_text)
        _safe_set(ws.cell(row=r, column=C.COL_CAMPUS), C.CONST_CAMPUS)
        _safe_set(ws.cell(row=r, column=C.COL_STATUT), C.CONST_STATUT)
        _safe_set(ws.cell(row=r, column=C.COL_DEMANDEUR), entry.demandeur)
        _safe_set(ws.cell(row=r, column=C.COL_LBUDG), entry.ligne_budgetaire)
        _safe_set(ws.cell(row=r, column=C.COL_ENTITE), entry.entite)

        if entry.status_note:
            nom_c = ws.cell(row=r, column=C.COL_NOM)
            if not isinstance(nom_c, MergedCell):
                nom_c.comment = Comment(entry.status_note, "Sync")

        for day, persons in entry.occupancy.items():
            col = lay.col_for(day)
            if col is None:
                continue
            cell = ws.cell(row=r, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = persons
            hexfill = C.color_for_persons(persons)
            if hexfill:
                cell.fill = PatternFill(start_color="FF" + hexfill,
                                        end_color="FF" + hexfill, fill_type="solid")
            note = entry.notes.get(day)
            if note:
                cell.comment = Comment(note, "Sync")

        nom_cell = ws.cell(row=r, column=C.COL_NOM)
        if nom_cell.value and len(str(nom_cell.value)) > 30:
            al = copy.copy(nom_cell.alignment)
            al.wrap_text = True
            nom_cell.alignment = al

        # grouper : colorer TOUTES les cellules d'info avec la couleur des jours
        # (4-9 -> orange clair, 10+ -> orange), identique aux cellules de jours.
        group_persons = max(entry.occupancy.values(), default=0)
        if group_persons > C.DEFAULT_CONFIG["group_threshold"]:
            hexfill = C.color_for_persons(group_persons)
            if hexfill:
                for c in (C.COL_NOM, C.COL_REF, C.COL_CAMPUS, C.COL_STATUT,
                          C.COL_DEMANDEUR, C.COL_LBUDG, C.COL_ENTITE):
                    cell = ws.cell(row=r, column=c)
                    if isinstance(cell, MergedCell):
                        continue
                    cell.fill = PatternFill(start_color="FF" + hexfill,
                                            end_color="FF" + hexfill,
                                            fill_type="solid")

        # toutes les lignes de données à la même hauteur
        if C.ROW_HEIGHT:
            ws.row_dimensions[r].height = C.ROW_HEIGHT

    def save(self, path=None):
        out = str(path or self.path)
        self.wb.save(out)
        return out

    # -- overstay (no check-out) -------------------------------------------
    def mark_overstay(self, sheet_title: str, row: int,
                      start_day: int, end_day: int, persons: int, note: str):
        """Fill extra days [start_day..end_day] with `persons`, red fill + note.

        Used by the Vérifier Planning check when a guest's real departure
        (Départ réel) is after the expected departure (Date de départ)."""
        ws = None
        for w in self.wb.worksheets:
            if w.title == sheet_title:
                ws = w
                break
        if ws is None:
            return
        lay = self.layout(ws)
        end_day = min(end_day, lay.last_day or end_day)
        for day in range(start_day, end_day + 1):
            col = lay.col_for(day)
            if col is None:
                continue
            cell = ws.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = persons
            cell.fill = PatternFill(start_color="FF" + C.COLOR_OVERSTAY,
                                    end_color="FF" + C.COLOR_OVERSTAY,
                                    fill_type="solid")
        col0 = lay.col_for(start_day)
        if col0:
            c0 = ws.cell(row=row, column=col0)
            if not isinstance(c0, MergedCell):
                c0.comment = Comment(note, "Sync")
